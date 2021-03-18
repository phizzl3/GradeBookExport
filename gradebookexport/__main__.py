import csv

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder

import dropfile


class Grades:

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def set_column_width(self, column_width=10):
        """
        Adjust all columns to the same specified width.
        """
        dim_holder = DimensionHolder(worksheet=xl.ws)
        for col in range(xl.ws.min_column, xl.ws.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(
                xl.ws, min=col, max=col, width=column_width)
        xl.ws.column_dimensions = dim_holder

    def remove_ungraded_columns(self):
        """
        Remove any columns that contain "Ungraded" content...multi-pass
        because I'm lazy.
        """
        for i in range(5):
            for col in xl.ws.iter_cols():
                for cell in col:
                    if cell.value:
                        if 'Ungraded' in cell.value:
                            xl.ws.delete_cols(cell.column, 1)

    def copy_csv_data(self, incsv):
        """
        Copy all values from csv file to target Excel Worksheet.

        Args:
            incsv (pathlib.Path): Path object representing a csv file.
        """
        try:
            with open(incsv, 'r') as f:
                reader = csv.reader(f)
                [self.ws.append(row) for row in reader]

        except Exception as e:
            print(f"\nError - copy_csv_data: {e}")
            input("[ENTER] to continue...")


if __name__ == '__main__':
    while True:
        print('\n Drag "gradebook-export.csv" to this window and press ENTER.')
        _csv = dropfile.get()
        xl = Grades()
        xl.copy_csv_data(_csv)
        xl.set_column_width()
        xl.remove_ungraded_columns()
        print(f'\n "{_csv.stem}.xlsx" will be saved to the same folder.')
        xl.wb.save(f'{_csv.parent}/{_csv.stem}.xlsx')
        input('\n Press ENTER to work on the next file...')
